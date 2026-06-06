"""
Robokidy Innovative Centre - Billing Software v5.0
- Edit/Delete: Students, Invoices, Payments, Course Plans, Finance entries
- Course Plans with editable fees (incl. Yearly All-Courses)
- Finance Tracker: Income & Expenses with categories, monthly summary
- Logo fix for .exe: embedded as base64 from exe folder
"""
from __future__ import annotations
import base64, io, sys, time, urllib.request, webbrowser
from datetime import datetime
from html import escape
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog

from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PRIMARY   = "00186D"
SECONDARY = "22BF96"
ACCENT    = "FF6B35"
DANGER    = "DC2626"
SUCCESS   = "16A34A"
BG_LIGHT  = "#F0F4FF"
BG_WHITE  = "#FFFFFF"

def _base_dir() -> Path:
    if getattr(sys,"frozen",False): return Path(sys.executable).parent
    return Path(__file__).parent

BASE_DIR    = _base_dir()
DATA_FILE   = BASE_DIR / "robokidy_billing_data.xlsx"
INVOICE_DIR = BASE_DIR / "invoices"
LOGO_URL    = "https://podu.pics/eR_8Q0TGcG"
LOGO_FILE   = BASE_DIR / "robokidy_logo.png"

COURSES = [
    "Python","Lego","Scratch","Robotics","Arduino","IoT","AI/ML",
    "Web Dev","App Dev","3D Printing","Drone Tech","Electronics",
    "Yearly All-Courses","Other"
]
COURSE_CODES = {
    "Python":"PY","Lego":"LG","Scratch":"SC","Robotics":"RB","Arduino":"AR","IoT":"IT",
    "AI/ML":"AI","Web Dev":"WD","App Dev":"AD","3D Printing":"3D",
    "Drone Tech":"DR","Electronics":"EL","Yearly All-Courses":"YR","Other":"OT"
}
LEVELS      = ["Level 1","Level 2","Level 3","Level 4","Advanced","Beginner","All Levels"]
BATCHES     = ["Morning","Afternoon","Evening","Weekend","Online"]
PAY_MODES   = ["Cash","UPI","Bank Transfer","Cheque","Card","Online"]
DURATIONS   = [f"{i} Month" for i in range(1,13)]
INC_CATS    = ["Course Fee","Registration Fee","Material Fee","Workshop","Donation","Other Income"]
EXP_CATS    = ["Rent","Electricity","Salaries","Equipment","Materials","Marketing",
               "Internet","Maintenance","Software","Miscellaneous"]
MONTHS      = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

STUDENT_HEADERS = ["Student ID","Student Name","Parent Name","Phone","Email",
                   "Grade/Age","Course","Course Code","Level","Batch",
                   "Joining Date","Status","Address","Notes"]
INVOICE_HEADERS = ["Invoice No","Date","Student ID","Student Name","Course","Level",
                   "Batch","Package/Fee","Discount Type","Discount Value","Discount Amount",
                   "Tax Type","Tax Value","Tax Amount","Total Amount",
                   "Paid Amount","Balance","Payment Status","Payment Mode","Due Date","Notes"]
PAYMENT_HEADERS = ["Receipt No","Date","Invoice No","Student ID","Student Name",
                   "Payment Mode","Paid Amount","Received By","Notes",
                   "Total Fees","Fees Paid","Balance Fees","Receipt Paid","Payment Status"]
PLAN_HEADERS    = ["Plan ID","Course","Level","Duration","Fee","Description","Active"]
COURSE_HEADERS  = ["Course","Course Code","Active"]
FINANCE_HEADERS = ["Entry ID","Date","Type","Category","Amount","Description","Payment Mode","Reference","Notes"]
SETTINGS_HEADERS= ["Key","Value"]

def money(v) -> float:
    try: return round(float(v),2)
    except: return 0.0

def safe_save(wb, path:Path):
    for i in range(3):
        try: wb.save(path); return
        except PermissionError:
            if i<2: time.sleep(0.5)
            else: raise PermissionError(
                f"File is open in Excel. Please close:\n{path.resolve()}\nThen try again.")

def logo_path() -> Path | None:
    bundled = Path(getattr(sys, "_MEIPASS", BASE_DIR)) / LOGO_FILE.name
    for candidate in (LOGO_FILE, bundled):
        if candidate.exists():
            return candidate
    try:
        req = urllib.request.Request(LOGO_URL, headers={"User-Agent":"Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=12) as r:
            raw = r.read()
        img = Image.open(io.BytesIO(raw))
        img.thumbnail((900, 900))
        img.save(LOGO_FILE, "PNG")
        return LOGO_FILE
    except Exception:
        return None

def logo_data_uri() -> str:
    path = logo_path()
    if not path:
        return LOGO_URL
    try:
        data = base64.b64encode(path.read_bytes()).decode("ascii")
        return f"data:image/png;base64,{data}"
    except Exception:
        return LOGO_URL

# ═══════════════════════════════════════════════════════════════════════════════
class ExcelStore:
    def __init__(self, path:Path=DATA_FILE):
        self.path = path; self._ensure()

    def _ensure(self):
        if self.path.exists():
            self._ensure_schema()
            return
        wb = Workbook(); ws = wb.active; ws.title = "Dashboard"
        for name,hdrs in [("Students",STUDENT_HEADERS),("Invoices",INVOICE_HEADERS),
                          ("Payments",PAYMENT_HEADERS),("CoursePlans",PLAN_HEADERS),
                          ("Courses",COURSE_HEADERS),("Finance",FINANCE_HEADERS),
                          ("Settings",SETTINGS_HEADERS)]:
            w = wb.create_sheet(name); w.append(hdrs); self._sh(w)
        ws.append(["ROBOKIDY INNOVATIVE CENTRE - BILLING DASHBOARD"])
        s = wb["Settings"]
        for k,v in [("Company Name","Robokidy Innovative Centre"),
                    ("Address","91/3 Ramachandra Nagar, Kallakurichi"),
                    ("Phone","+91 8300967241"),("Email","info@robokidy.com"),
                    ("Website","www.robokidy.com"),("GSTIN","33AAOCR3798M1Z1"),
                    ("Invoice Prefix","RKI-INV"),("Receipt Prefix","RKI-RCP")]:
            s.append([k,v])
        cp = wb["CoursePlans"]
        for row in [
            ("PLN-001","Python","Level 1","1 Month",2500,"Python basics","Yes"),
            ("PLN-002","Python","Level 2","2 Month",3000,"Python advanced","Yes"),
            ("PLN-003","Lego","Level 1","1 Month",2500,"Lego basics","Yes"),
            ("PLN-004","Robotics","Level 1","1 Month",3500,"Robotics basics","Yes"),
            ("PLN-005","Arduino","Level 1","1 Month",3000,"Arduino basics","Yes"),
            ("PLN-006","Scratch","Level 1","1 Month",2000,"Scratch basics","Yes"),
            ("PLN-007","AI/ML","Level 1","1 Month",4000,"AI/ML monthly","Yes"),
            ("PLN-008","IoT","Level 1","1 Month",3500,"IoT monthly","Yes"),
            ("PLN-009","Web Dev","Level 1","1 Month",3000,"Web Dev monthly","Yes"),
            ("PLN-010","Yearly All-Courses","All Levels","12 Month",25000,"All courses yearly plan","Yes"),
        ]: cp.append(row)
        cs = wb["Courses"]
        for course in COURSES:
            if course!="Other":
                cs.append([course,COURSE_CODES.get(course,course[:2].upper()),"Yes"])
        safe_save(wb,self.path)

    def _ensure_schema(self):
        wb=self._wb(); changed=False
        for name,hdrs in [("Students",STUDENT_HEADERS),("Invoices",INVOICE_HEADERS),
                          ("Payments",PAYMENT_HEADERS),("CoursePlans",PLAN_HEADERS),
                          ("Courses",COURSE_HEADERS),("Finance",FINANCE_HEADERS),
                          ("Settings",SETTINGS_HEADERS)]:
            if name not in wb.sheetnames:
                ws=wb.create_sheet(name); ws.append(hdrs); changed=True
            ws=wb[name]
            current=[c.value for c in ws[1]]
            for h in hdrs:
                if h not in current:
                    ws.cell(1,ws.max_column+1).value=h
                    current.append(h); changed=True
            self._sh(ws)
        if changed:
            safe_save(wb,self.path)

    def _wb(self): return load_workbook(self.path)

    @staticmethod
    def _sh(ws):
        fill=PatternFill("solid",fgColor=PRIMARY); thin=Side(style="thin",color="D9E2F3")
        for c in ws[1]:
            c.fill=fill; c.font=Font(color="FFFFFF",bold=True)
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.border=Border(top=thin,left=thin,right=thin,bottom=thin)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment=Alignment(vertical="center",wrap_text=True)
                c.border=Border(top=thin,left=thin,right=thin,bottom=thin)
        ws.freeze_panes="A2"
        for i,col in enumerate(ws.columns,1):
            values=[str(c.value) for c in col if c.value is not None]
            width=min(max([len(v) for v in values]+[12])+2,32)
            ws.column_dimensions[get_column_letter(i)].width=width

    def rows(self,sheet): 
        wb=self._wb(); ws=wb[sheet]; hdrs=[c.value for c in ws[1]]
        return [dict(zip(hdrs,r)) for r in ws.iter_rows(min_row=2,values_only=True)
                if any(v is not None for v in r)]

    def _append(self,sheet,vals):
        wb=self._wb(); ws=wb[sheet]; ws.append(vals); self._sh(ws); safe_save(wb,self.path)

    def _update(self,sheet,id_col,id_val,kv:dict):
        wb=self._wb(); ws=wb[sheet]; hdrs=[c.value for c in ws[1]]
        ic=hdrs.index(id_col)+1
        for r in range(2,ws.max_row+1):
            if str(ws.cell(r,ic).value)==str(id_val):
                for col,val in kv.items():
                    if col in hdrs: ws.cell(r,hdrs.index(col)+1).value=val
                break
        safe_save(wb,self.path)

    def _delete(self,sheet,id_col,id_val):
        wb=self._wb(); ws=wb[sheet]; hdrs=[c.value for c in ws[1]]
        ic=hdrs.index(id_col)+1
        for r in range(2,ws.max_row+1):
            if str(ws.cell(r,ic).value)==str(id_val): ws.delete_rows(r); break
        safe_save(wb,self.path)

    def settings(self): return {r["Key"]:r["Value"] for r in self.rows("Settings")}

    def course_values(self):
        vals=list(COURSES)
        for r in self.rows("Courses"):
            if str(r.get("Active","Yes"))!="No":
                c=str(r.get("Course") or "").strip()
                if c and c not in vals:
                    vals.insert(max(len(vals)-1,0),c) if "Other" in vals else vals.append(c)
        for sheet in ("Students","CoursePlans","Invoices"):
            for r in self.rows(sheet):
                c=str(r.get("Course") or "").strip()
                if c and c not in vals:
                    vals.insert(max(len(vals)-1,0),c) if "Other" in vals else vals.append(c)
        return vals

    def add_course(self,name):
        name=str(name or "").strip()
        if not name: raise ValueError("Course name required")
        existing=[str(r.get("Course") or "").strip().lower() for r in self.rows("Courses")]
        if name.lower() in existing or name in COURSES:
            raise ValueError("Course already exists")
        code="".join(ch for ch in name.upper() if ch.isalnum())[:2] or "CR"
        self._append("Courses",[name,code,"Yes"])
        return name

    def update_dashboard(self):
        wb=self._wb(); ws=wb["Dashboard"]; ws.delete_rows(1,ws.max_row)
        inv=self.rows("Invoices"); stu=self.rows("Students"); pay=self.rows("Payments")
        fin=self.rows("Finance")
        income =sum(money(r.get("Amount")) for r in fin if r.get("Type")=="Income")
        expense=sum(money(r.get("Amount")) for r in fin if r.get("Type")=="Expense")
        ws.append(["ROBOKIDY INNOVATIVE CENTRE - BILLING DASHBOARD"])
        ws.append(["Last Updated",datetime.now().strftime("%d-%m-%Y %I:%M %p")]); ws.append([])
        ws.append(["Total Students",len(stu)]); ws.append(["Total Invoices",len(inv)])
        ws.append(["Total Receipts",len(pay)])
        ws.append(["Total Billing",sum(money(r.get("Total Amount")) for r in inv)])
        ws.append(["Total Received",sum(money(r.get("Paid Amount")) for r in inv)])
        ws.append(["Total Balance",sum(money(r.get("Balance")) for r in inv)])
        ws.append([]); ws.append(["Total Income",income])
        ws.append(["Total Expenses",expense]); ws.append(["Net Profit",income-expense])
        ws["A1"].font=Font(size=16,bold=True,color=PRIMARY)
        ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=25
        safe_save(wb,self.path)

    # ── ID generators ──────────────────────────────────────────────────────────
    def _next_id(self,sheet,id_col,prefix,fmt="{p}{n:04d}",year=True):
        p = f"{prefix}-{datetime.now().year}-" if year else f"{prefix}-"
        existing=[r.get(id_col,"") for r in self.rows(sheet)]
        nums=[int(str(v).split("-")[-1]) for v in existing
              if str(v).startswith(p) and str(v).split("-")[-1].isdigit()]
        n = max(nums)+1 if nums else 1
        return fmt.format(p=p,n=n)

    def next_student_id(self,course,level):
        code=COURSE_CODES.get(course,course[:2].upper())
        lvl=level.replace("Level ","L").replace("All Levels","ALL").replace(" ","").upper()
        p=f"RK-{code}-{lvl}-{datetime.now().year}-"
        existing=[r.get("Student ID","") for r in self.rows("Students")]
        nums=[int(str(v).split("-")[-1]) for v in existing
              if str(v).startswith(p) and str(v).split("-")[-1].isdigit()]
        return f"{p}{(max(nums)+1 if nums else 1):03d}"

    def next_inv(self):  return self._next_id("Invoices","Invoice No",self.settings().get("Invoice Prefix","RKI-INV"))
    def next_rcp(self):  return self._next_id("Payments","Receipt No",self.settings().get("Receipt Prefix","RKI-RCP"))
    def next_plan(self): return self._next_id("CoursePlans","Plan ID","PLN",year=False)
    def next_fin(self):  return self._next_id("Finance","Entry ID","FIN")

    # ── Students ───────────────────────────────────────────────────────────────
    def add_student(self,d):
        sid=self.next_student_id(d["course"],d["level"])
        code=COURSE_CODES.get(d["course"],d["course"][:2].upper())
        self._append("Students",[sid,d["name"],d["parent"],d["phone"],d["email"],
            d["grade"],d["course"],code,d["level"],d["batch"],
            d["joining"],"Active",d["address"],d["notes"]])
        self.update_dashboard(); return sid

    def edit_student(self,sid,d):
        code=COURSE_CODES.get(d["course"],d["course"][:2].upper())
        self._update("Students","Student ID",sid,{
            "Student Name":d["name"],"Parent Name":d["parent"],"Phone":d["phone"],
            "Email":d["email"],"Grade/Age":d["grade"],"Course":d["course"],
            "Course Code":code,"Level":d["level"],"Batch":d["batch"],
            "Joining Date":d["joining"],"Status":d["status"],
            "Address":d["address"],"Notes":d["notes"]})
        self.update_dashboard()

    def delete_student(self,sid):
        self._delete("Students","Student ID",sid); self.update_dashboard()

    # ── Invoices ───────────────────────────────────────────────────────────────
    def create_invoice(self,d):
        inv_no=self.next_inv(); fee=money(d["fee"])
        da=round(fee*money(d["dv"])/100,2) if d["dt"]=="%" else money(d["dv"])
        sub=fee-da
        ta=round(sub*money(d["tv"])/100,2) if d["tt"]=="%" else money(d["tv"])
        total=round(sub+ta,2); paid=0.0; bal=total; status="Unpaid"
        self._append("Invoices",[inv_no,d["date"],d["sid"],d["sname"],d["course"],d["level"],
            d["batch"],fee,d["dt"],money(d["dv"]),da,d["tt"],money(d["tv"]),ta,
            total,paid,bal,status,"",d["due"],d["notes"]])
        self.update_dashboard(); return inv_no

    @staticmethod
    def _payment_status(total,paid):
        bal=round(money(total)-money(paid),2)
        if paid<=0:
            return "Unpaid"
        return "Paid" if bal<=0 else "Partial"

    def _sync_invoice_payment_state(self,wb,inv_no):
        inv_no=str(inv_no or "")
        if not inv_no:
            return
        iw=wb["Invoices"]; ih=[c.value for c in iw[1]]
        pw=wb["Payments"]; ph=[c.value for c in pw[1]]
        required=("Invoice No","Paid Amount","Balance","Total Amount","Payment Status","Payment Mode")
        if any(h not in ih for h in required):
            return
        if any(h not in ph for h in ("Invoice No","Paid Amount","Payment Mode")):
            return
        inv_row=None
        inv_col=ih.index("Invoice No")+1
        for r in range(2,iw.max_row+1):
            if str(iw.cell(r,inv_col).value)==inv_no:
                inv_row=r
                break
        if not inv_row:
            return

        total=money(iw.cell(inv_row,ih.index("Total Amount")+1).value)
        running=0.0; last_mode=""
        p_inv_col=ph.index("Invoice No")+1
        p_paid_col=ph.index("Paid Amount")+1
        p_mode_col=ph.index("Payment Mode")+1
        for r in range(2,pw.max_row+1):
            if str(pw.cell(r,p_inv_col).value)!=inv_no:
                continue
            receipt_paid=money(pw.cell(r,p_paid_col).value)
            running=round(running+receipt_paid,2)
            balance=round(total-running,2)
            mode=str(pw.cell(r,p_mode_col).value or "")
            if mode:
                last_mode=mode
            updates={
                "Total Fees":total,
                "Fees Paid":running,
                "Balance Fees":balance,
                "Receipt Paid":receipt_paid,
                "Payment Status":self._payment_status(total,running),
            }
            for col,val in updates.items():
                if col in ph:
                    pw.cell(r,ph.index(col)+1).value=val

        balance=round(total-running,2)
        iw.cell(inv_row,ih.index("Paid Amount")+1).value=running
        iw.cell(inv_row,ih.index("Balance")+1).value=balance
        iw.cell(inv_row,ih.index("Payment Status")+1).value=self._payment_status(total,running)
        iw.cell(inv_row,ih.index("Payment Mode")+1).value=last_mode if running>0 else ""

    def _delete_finance_by_reference(self,wb,reference):
        reference=str(reference or "")
        if not reference or "Finance" not in wb.sheetnames:
            return
        ws=wb["Finance"]; hdrs=[c.value for c in ws[1]]
        if "Reference" not in hdrs:
            return
        rc=hdrs.index("Reference")+1
        for r in range(ws.max_row,1,-1):
            if str(ws.cell(r,rc).value or "")==reference:
                ws.delete_rows(r)

    def delete_invoice(self,inv_no):
        inv_no=str(inv_no or "")
        wb=self._wb()
        if "Payments" in wb.sheetnames:
            pw=wb["Payments"]; ph=[c.value for c in pw[1]]
            if "Invoice No" in ph:
                ic=ph.index("Invoice No")+1
                rc=ph.index("Receipt No")+1 if "Receipt No" in ph else None
                for r in range(pw.max_row,1,-1):
                    if str(pw.cell(r,ic).value or "")==inv_no:
                        if rc:
                            self._delete_finance_by_reference(wb,pw.cell(r,rc).value)
                        pw.delete_rows(r)
        ws=wb["Invoices"]; hdrs=[c.value for c in ws[1]]
        ic=hdrs.index("Invoice No")+1
        for r in range(2,ws.max_row+1):
            if str(ws.cell(r,ic).value)==inv_no:
                ws.delete_rows(r)
                break
        for sheet in ("Invoices","Payments","Finance"):
            if sheet in wb.sheetnames:
                self._sh(wb[sheet])
        safe_save(wb,self.path)
        self.update_dashboard()

    def get_invoice(self,inv_no):
        for r in self.rows("Invoices"):
            if str(r.get("Invoice No"))==str(inv_no): return r
        return None

    def find_invoice_for_payment(self,key):
        key=str(key).strip()
        inv=self.get_invoice(key)
        if inv: return inv
        matches=[r for r in self.rows("Invoices") if str(r.get("Student ID"))==key]
        if not matches: return None
        pending=[r for r in matches if money(r.get("Balance"))>0]
        return (pending or matches)[-1]

    # ── Payments ───────────────────────────────────────────────────────────────
    def add_payment(self,d):
        r=self._save_pay(d["inv"],d["sid"],d["sname"],d["mode"],d["paid"],d["rcvd"],d["notes"])
        self.update_dashboard(); return r

    def _save_pay(self,inv_no,sid,sname,mode,paid,rcvd,notes,update=True):
        rno=self.next_rcp()
        pamt=money(paid); pay_date=datetime.now().strftime("%d-%m-%Y")
        if update:
            wb=self._wb(); ws=wb["Invoices"]; hdrs=[c.value for c in ws[1]]
            ic=hdrs.index("Invoice No")+1; pc=hdrs.index("Paid Amount")+1
            bc=hdrs.index("Balance")+1;    tc=hdrs.index("Total Amount")+1
            sc=hdrs.index("Payment Status")+1; mc=hdrs.index("Payment Mode")+1
            for r in range(2,ws.max_row+1):
                if ws.cell(r,ic).value==inv_no:
                    np2=money(ws.cell(r,pc).value)+pamt
                    tot=money(ws.cell(r,tc).value); nb=round(tot-np2,2)
                    ws.cell(r,pc).value=np2; ws.cell(r,bc).value=nb
                    ws.cell(r,sc).value="Paid" if nb<=0 else "Partial"
                    ws.cell(r,mc).value=mode
                    break
            safe_save(wb,self.path)
        inv=self.get_invoice(inv_no) or {}
        total=money(inv.get("Total Amount"))
        paid_after=money(inv.get("Paid Amount"))
        balance=money(inv.get("Balance"))
        status=inv.get("Payment Status","")
        self._append("Payments",[rno,pay_date,inv_no,sid,sname,mode,pamt,rcvd,notes,
            total,paid_after,balance,pamt,status])
        self._append("Finance",[self.next_fin(),pay_date,"Income","Course Fee",pamt,
            f"Fee payment from {sname}",mode,rno,notes])
        return rno

    def delete_payment(self,rno):
        rno=str(rno or "")
        wb=self._wb()
        pw=wb["Payments"]; ph=[c.value for c in pw[1]]
        if "Receipt No" not in ph:
            return
        rc=ph.index("Receipt No")+1
        inv_no=""
        for r in range(2,pw.max_row+1):
            if str(pw.cell(r,rc).value)==rno:
                if "Invoice No" in ph:
                    inv_no=str(pw.cell(r,ph.index("Invoice No")+1).value or "")
                pw.delete_rows(r)
                break
        self._delete_finance_by_reference(wb,rno)
        if inv_no:
            self._sync_invoice_payment_state(wb,inv_no)
        for sheet in ("Invoices","Payments","Finance"):
            if sheet in wb.sheetnames:
                self._sh(wb[sheet])
        safe_save(wb,self.path)
        self.update_dashboard()

    def get_payment(self,rno):
        for r in self.rows("Payments"):
            if str(r.get("Receipt No"))==str(rno): return r
        return None

    # ── Course Plans ───────────────────────────────────────────────────────────
    def add_plan(self,d):
        pid=self.next_plan()
        self._append("CoursePlans",[pid,d["course"],d["level"],d["duration"],
            money(d["fee"]),d["desc"],d["active"]])
        return pid

    def edit_plan(self,pid,d):
        self._update("CoursePlans","Plan ID",pid,{
            "Course":d["course"],"Level":d["level"],"Duration":d["duration"],
            "Fee":money(d["fee"]),"Description":d["desc"],"Active":d["active"]})

    def delete_plan(self,pid):
        self._delete("CoursePlans","Plan ID",pid)

    # ── Finance ────────────────────────────────────────────────────────────────
    def add_finance(self,d):
        eid=self.next_fin()
        self._append("Finance",[eid,d["date"],d["type"],d["cat"],money(d["amount"]),
            d["desc"],d["mode"],d["ref"],d["notes"]])
        self.update_dashboard(); return eid

    def edit_finance(self,eid,d):
        self._update("Finance","Entry ID",eid,{
            "Date":d["date"],"Type":d["type"],"Category":d["cat"],
            "Amount":money(d["amount"]),"Description":d["desc"],
            "Payment Mode":d["mode"],"Reference":d["ref"],"Notes":d["notes"]})
        self.update_dashboard()

    def delete_finance(self,eid):
        self._delete("Finance","Entry ID",eid); self.update_dashboard()

# ── HTML Invoice ───────────────────────────────────────────────────────────────
def make_html(company,inv,receipt_no=None,payment=None):
    logo_html=(f'<img src="{escape(logo_data_uri(), quote=True)}" alt="Robokidy logo" '
               f'style="width:168px;height:58px;object-fit:contain;background:#fff;'
               f'border-radius:8px;padding:6px;box-shadow:0 4px 14px rgba(0,0,0,.16)">')
    is_r=receipt_no is not None
    badge="RECEIPT" if is_r else "INVOICE"
    bcol=SECONDARY if is_r else "ffffff"
    def row(l,v):
        return (f'<tr><td style="padding:10px 16px;background:#f4f6ff;font-weight:600;'
                f'color:#444;width:44%;border-bottom:1px solid #e0e6f8;">{l}</td>'
                f'<td style="padding:10px 16px;border-bottom:1px solid #e0e6f8;">{v}</td></tr>')
    da=money(inv.get("Discount Amount",0)); ta=money(inv.get("Tax Amount",0))
    dv=money(inv.get("Discount Value",0)); tv=money(inv.get("Tax Value",0))
    dt=inv.get("Discount Type","");        tt=inv.get("Tax Type","")
    disc=f"₹ {da:,.2f}"+( f" ({dv}%)" if dt=="%" else "")
    tax =f"₹ {ta:,.2f}"+( f" ({tv}% GST)" if tt=="%" else " (GST)")
    st=(payment or inv).get("Payment Status","")
    sc={"Paid":"#16a34a","Partial":"#d97706","Unpaid":"#dc2626"}.get(st,"#555")
    g=company.get("GSTIN","")
    gline=f'<div style="font-size:11px;color:#b8c8ff;margin-top:2px;">GSTIN: {g}</div>' if g else ""
    gst_details=f'<p><strong>GSTIN:</strong> {g}</p>' if g else ""
    doc_date=(payment or inv).get("Date","")
    receipt_paid=money((payment or {}).get("Receipt Paid") or (payment or {}).get("Paid Amount"))
    total_fees=money((payment or {}).get("Total Fees") or inv.get("Total Amount"))
    fees_paid=money((payment or {}).get("Fees Paid") or inv.get("Paid Amount"))
    balance_fees=money((payment or {}).get("Balance Fees") or inv.get("Balance"))
    title_no=receipt_no if is_r else inv.get("Invoice No","")
    return f"""<!doctype html><html><head><meta charset="utf-8">
<title>{badge} {title_no}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:Arial,Helvetica,sans-serif;background:#eef2ff;color:#222;-webkit-print-color-adjust:exact;print-color-adjust:exact}}
.pg{{max-width:800px;margin:24px auto;background:#fff;border-radius:14px;box-shadow:0 6px 32px rgba(0,24,109,.15);overflow:hidden}}
.hd{{background:linear-gradient(135deg,#{PRIMARY} 0%,#0a2d9c 100%);padding:24px 32px;display:flex;justify-content:space-between;align-items:center}}
.badge{{background:#{bcol};color:{'#'+PRIMARY if not is_r else '#fff'};font-weight:700;font-size:14px;padding:9px 22px;border-radius:24px;letter-spacing:2px}}
.body{{padding:28px 32px}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:24px}}
.box{{background:#f4f6ff;border:1.5px solid #dde3f8;border-radius:10px;padding:14px 18px}}
.box h4{{font-size:10px;text-transform:uppercase;letter-spacing:1.5px;color:#{SECONDARY};margin-bottom:6px;font-weight:700}}
.box p{{font-size:12.5px;line-height:1.8}}.box strong{{color:#{PRIMARY}}}
table{{width:100%;border-collapse:collapse;border:1.5px solid #dde3f8;border-radius:10px;overflow:hidden}}
.tot td{{background:#{PRIMARY}!important;color:#fff!important;font-weight:700;font-size:15px;padding:13px 16px!important}}
.sb{{display:inline-block;padding:3px 12px;border-radius:16px;font-size:11px;font-weight:700;color:{sc};background:{sc}18;border:1.5px solid {sc}44}}
.footer{{margin-top:32px;display:flex;justify-content:space-between;padding-top:16px;border-top:2px dashed #dde3f8}}
.sig{{text-align:center}}.sl{{width:150px;border-top:1.5px solid #555;margin:0 auto 5px;padding-top:5px;font-size:11px;color:#666}}
.pb{{display:flex;gap:10px;margin-bottom:20px}}
.btn{{padding:9px 22px;border:none;border-radius:7px;font-weight:700;font-size:12px;cursor:pointer}}
.bp{{background:#{PRIMARY};color:#fff}}.bs{{background:#{SECONDARY};color:#fff}}
@page{{margin:12mm}}
@media print{{.pb{{display:none!important}}body{{background:#fff}}.pg{{margin:0;box-shadow:none;border-radius:0}}.hd,.tot td,.btn{{-webkit-print-color-adjust:exact;print-color-adjust:exact}}}}
</style></head><body>
<div class="pg">
<div class="hd">
  <div style="display:flex;align-items:center;gap:18px">
    {logo_html}
    <div>
      <div style="color:#a8b8ff;font-size:11.5px">{company.get('Address','')}</div>
      <div style="color:#a8b8ff;font-size:11.5px">{company.get('Phone','')} | {company.get('Email','')}</div>
      {gline}
    </div>
  </div>
  <div class="badge">{badge}</div>
</div>
<div class="body">
  <div class="pb">
    <button class="btn bp" onclick="window.print()">Print / Save PDF</button>
    <button class="btn bs" onclick="window.close()">Close</button>
  </div>
  <div class="grid">
    <div class="box"><h4>{'Receipt' if is_r else 'Invoice'} Details</h4>
      <p><strong>{'Receipt' if is_r else 'Invoice'} No:</strong> {receipt_no if is_r else inv.get('Invoice No','')}</p>
      <p><strong>Date:</strong> {doc_date}</p>
      {gst_details}
      {'<p><strong>Due Date:</strong> '+str(inv.get('Due Date',''))+'</p>' if not is_r else ''}
    </div>
    <div class="box"><h4>Student Details</h4>
      <p><strong>ID:</strong> {inv.get('Student ID','')}</p>
      <p><strong>Name:</strong> {inv.get('Student Name','')}</p>
      <p><strong>Course:</strong> {inv.get('Course','')} | {inv.get('Level','')}</p>
      <p><strong>Batch:</strong> {inv.get('Batch','')}</p>
    </div>
  </div>
  <table>
    {row("Course / Program",f"{inv.get('Course','')} — {inv.get('Level','')}")}
    {row("Package / Fee",f"₹ {money(inv.get('Package/Fee')):,.2f}")}
    {row("Discount",disc if da else "—")}
    {row("Tax (GST)",tax if ta else "—")}
    <tr class="tot"><td>TOTAL FEES</td><td>₹ {total_fees:,.2f}</td></tr>
    {row("Paid This Receipt",f"₹ {receipt_paid:,.2f}") if is_r else ""}
    {row("Fees Paid",f"₹ {fees_paid:,.2f}")}
    {row("Balance Fees",f"₹ {balance_fees:,.2f}")}
    {row("Payment Mode",(payment or inv).get('Payment Mode',''))}
    {row("Payment Status",f'<span class="sb">{st}</span>')}
  </table>
  {'<p style="margin-top:14px;font-size:12px;color:#666"><b>Notes:</b> '+str(inv.get('Notes',''))+'</p>' if inv.get('Notes') else ''}
  <div class="footer">
    <div class="sig"><div class="sl">Parent / Student Signature</div></div>
    <div style="text-align:center;font-size:11px;color:#888">
      <div>Thank you for choosing Robokidy!</div>
      <div style="color:#{SECONDARY};font-weight:700">{company.get('Website','')}</div>
    </div>
    <div class="sig"><div class="sl">Authorised Signature</div></div>
  </div>
</div></div></body></html>"""

# ═══════════════════════════════════════════════════════════════════════════════
# UI HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def sbtn(parent,text,cmd,bg=f"#{PRIMARY}",fg="white",px=14,py=7):
    return tk.Button(parent,text=text,command=cmd,bg=bg,fg=fg,
                     font=("Segoe UI",9,"bold"),relief="flat",cursor="hand2",
                     padx=px,pady=py,bd=0,activebackground=bg)

def slbl(parent,text):
    return tk.Label(parent,text=text,bg=BG_WHITE,font=("Segoe UI",11,"bold"),fg=f"#{PRIMARY}")

# ── Generic Edit Dialog ────────────────────────────────────────────────────────
class Dialog(tk.Toplevel):
    def __init__(self,parent,title,w=660,h=500):
        super().__init__(parent); self.title(title); self.grab_set()
        self.resizable(True,True); self.geometry(f"{w}x{h}"); self.configure(bg=BG_WHITE)
        self.result=None
        hdr=tk.Frame(self,bg=f"#{PRIMARY}",height=44); hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr,text=f"  {title}",bg=f"#{PRIMARY}",fg="white",
                 font=("Segoe UI",11,"bold")).pack(side="left",padx=12)
        self.body=tk.Frame(self,bg=BG_WHITE,padx=20,pady=16)
        self.body.pack(fill="both",expand=True)
        self._build()
        br=tk.Frame(self,bg="#f0f4ff",pady=10); br.pack(fill="x")
        sbtn(br,"💾  Save",self._save,bg=f"#{SECONDARY}").pack(side="left",padx=14)
        sbtn(br,"✕  Cancel",self.destroy,bg="#888").pack(side="left",padx=4)
        self.transient(parent); self.wait_window(self)

    def ef(self,lbl,row,col=0,w=28,v="",colspan=1):
        tk.Label(self.body,text=lbl,bg=BG_WHITE,fg="#444",
                 font=("Segoe UI",9)).grid(row=row,column=col,sticky="w",padx=(0,6),pady=5)
        e=ttk.Entry(self.body,width=w)
        e.grid(row=row,column=col+1,columnspan=colspan,sticky="ew",pady=5,padx=(0,12))
        if v: e.insert(0,str(v))
        return e

    def cb(self,lbl,vals,row,col=0,w=26,v=""):
        tk.Label(self.body,text=lbl,bg=BG_WHITE,fg="#444",
                 font=("Segoe UI",9)).grid(row=row,column=col,sticky="w",padx=(0,6),pady=5)
        c=ttk.Combobox(self.body,values=vals,state="readonly",width=w)
        c.grid(row=row,column=col+1,sticky="ew",pady=5,padx=(0,12))
        if v and v in vals: c.set(v)
        elif vals: c.current(0)
        return c

    def _build(self): pass
    def _save(self): pass


class StudentDlg(Dialog):
    def __init__(self,parent,data=None):
        self._d=data or {}
        self._courses=getattr(parent,"_course_values",COURSES)
        super().__init__(parent,"✏️ Edit Student" if data else "➕ Add Student",700,520)
    def _build(self):
        d=self._d; f=self.body; f.columnconfigure((1,3),weight=1)
        self.nm=self.ef("Student Name *",0,0,v=d.get("Student Name",""))
        self.pn=self.ef("Parent Name",0,2,v=d.get("Parent Name",""))
        self.ph=self.ef("Phone *",1,0,v=d.get("Phone",""))
        self.em=self.ef("Email",1,2,v=d.get("Email",""))
        self.gr=self.ef("Grade / Age",2,0,v=d.get("Grade/Age",""))
        self.ad=self.ef("Address",2,2,v=d.get("Address",""))
        self.co=self.cb("Course *",self._courses,3,0,v=d.get("Course",""))
        if d.get("Course") and d.get("Course") not in self._courses:
            self.co["values"]=list(self._courses)+[d.get("Course")]
            self.co.set(d.get("Course"))
        self.lv=self.cb("Level *",LEVELS,3,2,v=d.get("Level",""))
        self.ba=self.cb("Batch",BATCHES,4,0,v=d.get("Batch",""))
        self.st=self.cb("Status",["Active","Inactive","On Hold"],4,2,v=d.get("Status","Active"))
        self.jd=self.ef("Joining Date",5,0,v=d.get("Joining Date",datetime.now().strftime("%d-%m-%Y")))
        self.no=self.ef("Notes",5,2,v=d.get("Notes",""))
    def _save(self):
        if not self.nm.get().strip(): return messagebox.showerror("Missing","Name required",parent=self)
        self.result={"name":self.nm.get().strip(),"parent":self.pn.get().strip(),
            "phone":self.ph.get().strip(),"email":self.em.get().strip(),
            "grade":self.gr.get().strip(),"course":self.co.get(),"level":self.lv.get(),
            "batch":self.ba.get(),"status":self.st.get(),"joining":self.jd.get(),
            "address":self.ad.get().strip(),"notes":self.no.get().strip()}
        self.destroy()


class PlanDlg(Dialog):
    def __init__(self,parent,data=None):
        self._d=data or {}
        self._courses=getattr(parent,"_course_values",COURSES)
        super().__init__(parent,"✏️ Edit Plan" if data else "➕ Add Course Plan",560,420)
    def _build(self):
        d=self._d; f=self.body; f.columnconfigure(1,weight=1)
        self.co=self.cb("Course *",self._courses,0,v=d.get("Course",""))
        other_course="" if d.get("Course","") in self._courses else d.get("Course","")
        if other_course:
            self.co.set("Other")
        self.oc=self.ef("Other Course",1,v=other_course)
        self.lv=self.cb("Level",LEVELS,2,v=d.get("Level","Level 1"))
        self.du=self.cb("Duration",DURATIONS,3,v=d.get("Duration","1 Month"))
        self.fe=self.ef("Fee (₹) *",4,v=str(d.get("Fee","")))
        self.de=self.ef("Description",5,w=36,v=d.get("Description",""))
        self.ac=self.cb("Active",["Yes","No"],6,v=d.get("Active","Yes"))
    def _save(self):
        if not self.fe.get().strip(): return messagebox.showerror("Missing","Fee required",parent=self)
        course=self.oc.get().strip() if self.co.get()=="Other" else self.co.get().strip()
        if self.co.get()=="Other" and not course:
            return messagebox.showerror("Missing","Enter the other course name",parent=self)
        self.result={"course":course,"level":self.lv.get(),"duration":self.du.get(),
            "fee":self.fe.get(),"desc":self.de.get().strip(),"active":self.ac.get()}
        self.destroy()


class FinanceDlg(Dialog):
    def __init__(self,parent,data=None,default_type="Income"):
        self._d=data or {}; self._dt=default_type
        super().__init__(parent,"✏️ Edit Entry" if data else "➕ Add Finance Entry",600,460)
    def _build(self):
        d=self._d; f=self.body; f.columnconfigure((1,3),weight=1)
        self.tp=self.cb("Type *",["Income","Expense"],0,0,v=d.get("Type",self._dt))
        self.dt=self.ef("Date *",0,2,v=d.get("Date",datetime.now().strftime("%d-%m-%Y")))
        self._cat_var=tk.StringVar()
        tk.Label(f,text="Category *",bg=BG_WHITE,fg="#444",
                 font=("Segoe UI",9)).grid(row=1,column=0,sticky="w",padx=(0,6),pady=5)
        self.ca=ttk.Combobox(f,textvariable=self._cat_var,width=26,state="readonly")
        self.ca.grid(row=1,column=1,sticky="ew",pady=5,padx=(0,12))
        self.tp.bind("<<ComboboxSelected>>",self._set_cats)
        self._set_cats(init=d.get("Category",""))
        self.am=self.ef("Amount (₹) *",1,2,v=str(d.get("Amount","")))
        self.de=self.ef("Description",2,0,w=28,v=d.get("Description",""))
        self.mo=self.cb("Payment Mode",PAY_MODES,2,2,v=d.get("Payment Mode","Cash"))
        self.rf=self.ef("Reference",3,0,v=d.get("Reference",""))
        self.no=self.ef("Notes",3,2,v=d.get("Notes",""))

    def _set_cats(self,_=None,init=""):
        cats=INC_CATS if self.tp.get()=="Income" else EXP_CATS
        self.ca["values"]=cats
        cur=init or self._cat_var.get()
        if cur in cats: self.ca.set(cur)
        else: self.ca.current(0)

    def _save(self):
        if not self.am.get().strip(): return messagebox.showerror("Missing","Amount required",parent=self)
        self.result={"type":self.tp.get(),"date":self.dt.get(),"cat":self.ca.get(),
            "amount":self.am.get(),"desc":self.de.get().strip(),"mode":self.mo.get(),
            "ref":self.rf.get().strip(),"notes":self.no.get().strip()}
        self.destroy()

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN APP
# ═══════════════════════════════════════════════════════════════════════════════
class BillingApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Billing Software v5.0")
        self.geometry("1260x820"); self.minsize(980,640)
        self.store=ExcelStore(); self.configure(bg=BG_LIGHT)
        INVOICE_DIR.mkdir(exist_ok=True)
        self._course_values=list(COURSES)
        self._slookup={}; self._plan_map={}
        self._build(); self.refresh_all()

    # ── Shell ──────────────────────────────────────────────────────────────────
    def _build(self):
        tb=tk.Frame(self,bg=f"#{PRIMARY}",height=56); tb.pack(fill="x"); tb.pack_propagate(False)
        self._logo_img=self._load_logo_image((150,42))
        if self._logo_img:
            tk.Label(tb,image=self._logo_img,bg=f"#{PRIMARY}").pack(side="left",padx=(14,0))
        tk.Label(tb,text="Billing v5.0",bg=f"#{PRIMARY}",fg=f"#{SECONDARY}",
                 font=("Segoe UI",11,"bold")).pack(side="left",padx=14)
        self._clk=tk.Label(tb,bg=f"#{PRIMARY}",fg="#aabbff",font=("Segoe UI",10))
        self._clk.pack(side="right",padx=16); self._tick()
        s=ttk.Style(self); s.theme_use("clam")
        s.configure("TNotebook",background=BG_LIGHT,borderwidth=0)
        s.configure("TNotebook.Tab",background="#dde3f8",foreground="#1A1A2E",
                    padding=[14,9],font=("Segoe UI",9,"bold"))
        s.map("TNotebook.Tab",background=[("selected",f"#{SECONDARY}")],
              foreground=[("selected","white")])
        s.configure("Treeview",rowheight=27,font=("Segoe UI",9))
        s.configure("Treeview.Heading",font=("Segoe UI",9,"bold"),
                    background=f"#{PRIMARY}",foreground="white")
        s.map("Treeview",background=[("selected",f"#{SECONDARY}")])
        self.nb=ttk.Notebook(self); self.nb.pack(fill="both",expand=True,padx=10,pady=8)
        self._t_dash(); self._t_stu(); self._t_inv(); self._t_pay()
        self._t_plans(); self._t_finance(); self._t_reports()

    def _load_logo_image(self,size):
        path=logo_path()
        if not path:
            return None
        try:
            img=Image.open(path).convert("RGBA")
            img.thumbnail(size)
            canvas=Image.new("RGBA",size,(255,255,255,0))
            x=(size[0]-img.width)//2; y=(size[1]-img.height)//2
            canvas.paste(img,(x,y),img)
            return ImageTk.PhotoImage(canvas)
        except Exception:
            return None

    def _tick(self):
        self._clk.config(text=datetime.now().strftime("  %d %b %Y   %I:%M:%S %p  "))
        self.after(1000,self._tick)

    # ── Widget factories ───────────────────────────────────────────────────────
    def _ff(self,p,lbl,row,col,w=22,cs=1,val=""):
        tk.Label(p,text=lbl,bg=BG_WHITE,fg="#444",font=("Segoe UI",9)).grid(
            row=row,column=col,sticky="w",padx=(10,4),pady=4)
        e=ttk.Entry(p,width=w)
        e.grid(row=row,column=col+1,columnspan=cs,sticky="ew",padx=(0,10),pady=4)
        if val: e.insert(0,val)
        return e

    def _fc(self,p,lbl,vals,row,col,w=22,val=""):
        tk.Label(p,text=lbl,bg=BG_WHITE,fg="#444",font=("Segoe UI",9)).grid(
            row=row,column=col,sticky="w",padx=(10,4),pady=4)
        c=ttk.Combobox(p,values=vals,state="readonly",width=w)
        c.grid(row=row,column=col+1,sticky="ew",padx=(0,10),pady=4)
        if val: c.set(val)
        return c

    def _treef(self,parent,cols,h=10):
        fr=tk.Frame(parent,bg=BG_WHITE); fr.pack(fill="both",expand=True,padx=10,pady=(0,6))
        t=ttk.Treeview(fr,columns=cols,show="headings",height=h,selectmode="browse")
        vs=ttk.Scrollbar(fr,orient="vertical",command=t.yview)
        hs=ttk.Scrollbar(fr,orient="horizontal",command=t.xview)
        t.configure(yscrollcommand=vs.set,xscrollcommand=hs.set)
        for c in cols: t.heading(c,text=c); t.column(c,width=110,anchor="center",minwidth=60)
        t.grid(row=0,column=0,sticky="nsew"); vs.grid(row=0,column=1,sticky="ns")
        hs.grid(row=1,column=0,sticky="ew")
        fr.grid_rowconfigure(0,weight=1); fr.grid_columnconfigure(0,weight=1)
        t.tag_configure("odd",background="#f5f7ff")
        t.tag_configure("even",background=BG_WHITE)
        t.tag_configure("income",background="#f0fdf4")
        t.tag_configure("expense",background="#fff7f7")
        return t

    def _fill(self,tree,rows,hdrs):
        for i in tree.get_children(): tree.delete(i)
        for idx,r in enumerate(rows):
            tag="odd" if idx%2 else "even"
            tree.insert("","end",values=[r.get(h,"") for h in hdrs],tags=(tag,))

    def _fill_fin(self,tree,rows,hdrs):
        for i in tree.get_children(): tree.delete(i)
        for idx,r in enumerate(rows):
            tag="income" if r.get("Type")=="Income" else "expense"
            tree.insert("","end",values=[r.get(h,"") for h in hdrs],tags=(tag,))

    def _date_obj(self,ds):
        for fmt in ["%d-%m-%Y","%Y-%m-%d","%d/%m/%Y"]:
            try: return datetime.strptime(str(ds),fmt)
            except: pass
        return None

    def _date_ok(self,row_date,start,end):
        d=self._date_obj(row_date)
        if start and start!="All":
            sd=self._date_obj(start)
            if sd and (not d or d<sd): return False
        if end and end!="All":
            ed=self._date_obj(end)
            if ed and (not d or d>ed): return False
        return True

    def _student_pick_text(self,student):
        return f"{student.get('Student ID','')} - {student.get('Student Name','')}"

    def _student_id_from_pick(self,value):
        text=str(value or "").strip()
        if " - " in text:
            return text.split(" - ",1)[0].strip()
        return text

    def _filter_payment_students(self,_=None):
        if not hasattr(self,"pi_sid"): return
        if _ and getattr(_,"keysym","") in ("Return","Tab","Escape","Up","Down","Left","Right"):
            return
        q=self.pi_sid.get().strip().lower()
        vals=getattr(self,"_payment_student_values",[])
        if q:
            vals=[v for v in vals if v.lower().startswith(q) or
                  " - " in v and v.split(" - ",1)[1].lower().startswith(q) or
                  q in v.lower()]
        self.pi_sid["values"]=vals
        if vals:
            try: self.pi_sid.event_generate("<Down>")
            except: pass

    def _dates_for(self,rows,col):
        vals=sorted({str(r.get(col) or "") for r in rows if r.get(col)}, key=lambda x:self._date_obj(x) or datetime.min)
        return ["All"]+vals

    def _export_rows(self,title,rows,hdrs):
        if not rows: return messagebox.showwarning("Export","No rows to export")
        sel=filedialog.asksaveasfilename(defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],title=f"Save {title} Report",
            initialfile=f"{title.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            initialdir=str(BASE_DIR))
        if not sel: return
        wb=Workbook(); ws=wb.active; ws.title=title[:31]
        ws.append(hdrs)
        for r in rows: ws.append([r.get(h,"") for h in hdrs])
        ExcelStore._sh(ws); safe_save(wb,Path(sel))
        messagebox.showinfo("Exported",f"Report saved:\n{sel}")

    def _filter_bar(self,parent,fields,apply_cmd,clear_cmd,export_cmd=None):
        bar=tk.Frame(parent,bg=BG_WHITE); bar.pack(fill="x",padx=10,pady=(4,2))
        ctrls={}
        for label,vals,width in fields:
            tk.Label(bar,text=label,bg=BG_WHITE,font=("Segoe UI",9)).pack(side="left",padx=(6,2))
            cb=ttk.Combobox(bar,values=vals,state=("normal" if label in ("From","To") else "readonly"),width=width)
            cb.set("All"); cb.pack(side="left",padx=(0,6))
            ctrls[label]=cb
        sbtn(bar,"🔍 Filter",apply_cmd,bg=f"#{PRIMARY}",px=10,py=5).pack(side="left",padx=4)
        sbtn(bar,"🔄 Clear",clear_cmd,bg="#888",px=10,py=5).pack(side="left",padx=4)
        if export_cmd:
            sbtn(bar,"⬇ Export Excel",export_cmd,bg=f"#{SECONDARY}",px=10,py=5).pack(side="left",padx=4)
        return ctrls

    def _ff_frame(self,parent):
        f=tk.Frame(parent,bg=BG_WHITE,padx=8,pady=10); f.pack(fill="x",padx=10,pady=(8,2)); return f

    def _abar(self,parent,btns):
        bar=tk.Frame(parent,bg=BG_WHITE); bar.pack(fill="x",padx=10,pady=(2,6))
        for txt,cmd,bg in btns: sbtn(bar,txt,cmd,bg=bg).pack(side="left",padx=4,pady=3)
        return bar

    def _pe(self,msg): messagebox.showerror("File In Use ⚠️",msg)

    # ══════════════════════════════════════════════════════════════════════════
    # DASHBOARD
    # ══════════════════════════════════════════════════════════════════════════
    def _t_dash(self):
        f=tk.Frame(self.nb,bg=BG_WHITE); self.nb.add(f,text="📊  Dashboard")
        cr=tk.Frame(f,bg=BG_WHITE); cr.pack(fill="x",padx=20,pady=14)
        self._dc={}
        cards=[("Students",f"#{PRIMARY}"),("Invoices","#0a2d9c"),("Receipts",f"#{SECONDARY}"),
               ("Billing",f"#{ACCENT}"),("Received",f"#{SUCCESS}"),("Balance",f"#{DANGER}"),
               ("Income",f"#{SUCCESS}"),("Expenses",f"#{DANGER}"),("Net Profit","#7c3aed")]
        for key,col in cards:
            c=tk.Frame(cr,bg=BG_LIGHT,width=126,height=82)
            c.pack(side="left",padx=5,pady=4,expand=True,fill="both"); c.pack_propagate(False)
            tk.Label(c,text=key,bg=BG_LIGHT,font=("Segoe UI",8),fg="#666").pack(pady=(10,1))
            lbl=tk.Label(c,text="—",bg=BG_LIGHT,font=("Segoe UI",13,"bold"),fg=col)
            lbl.pack(); self._dc[key]=lbl
        bar=tk.Frame(f,bg=BG_WHITE); bar.pack(fill="x",padx=20,pady=6)
        sbtn(bar,"🔄 Refresh",self.refresh_all,bg=f"#{SECONDARY}").pack(side="left",padx=4)
        sbtn(bar,"📂 Open Excel",lambda:webbrowser.open(DATA_FILE.resolve().as_uri())).pack(side="left",padx=4)
        tk.Frame(f,bg="#dde3f8",height=2).pack(fill="x",padx=20,pady=8)
        tk.Label(f,text="  Recent Invoices",bg=BG_WHITE,font=("Segoe UI",10,"bold"),
                 fg=f"#{PRIMARY}").pack(anchor="w")
        sh=["Invoice No","Date","Student Name","Total Amount","Paid Amount","Balance","Payment Status"]
        self._dt=self._treef(f,sh,h=8)

    # ══════════════════════════════════════════════════════════════════════════
    # STUDENTS
    # ══════════════════════════════════════════════════════════════════════════
    def _t_stu(self):
        f=tk.Frame(self.nb,bg=BG_WHITE); self.nb.add(f,text="👤  Students")
        slbl(f,"  Add New Student").pack(anchor="w",padx=14,pady=(10,0))
        form=self._ff_frame(f); form.columnconfigure((1,3,5),weight=1)
        self.sn=self._ff(form,"Student Name *",0,0)
        self.sp=self._ff(form,"Parent Name",0,2)
        self.sph=self._ff(form,"Phone *",0,4)
        self.se=self._ff(form,"Email",1,0)
        self.sg=self._ff(form,"Grade / Age",1,2)
        self.sa=self._ff(form,"Address",1,4)
        self.sco=self._fc(form,"Course *",COURSES,2,0)
        self.slv=self._fc(form,"Level *",LEVELS,2,2)
        self.sba=self._fc(form,"Batch",BATCHES,2,4)
        self.sj=self._ff(form,"Joining Date",3,0,val=datetime.now().strftime("%d-%m-%Y"))
        self.sno=self._ff(form,"Notes",3,2,w=50,cs=3)
        self._sidlbl=tk.Label(form,text="ID preview: select Course & Level",
                              bg=BG_WHITE,font=("Segoe UI",8,"italic"),fg=f"#{SECONDARY}")
        self._sidlbl.grid(row=4,column=0,columnspan=4,sticky="w",padx=10)
        self.sco.bind("<<ComboboxSelected>>",self._prev_sid)
        self.slv.bind("<<ComboboxSelected>>",self._prev_sid)
        self._abar(f,[
            ("➕  Add Student",    self._add_stu,  f"#{SECONDARY}"),
            ("✏️  Edit Selected",  self._edit_stu, f"#{PRIMARY}"),
            ("🗑️  Delete Selected",self._del_stu,  f"#{DANGER}"),
            ("🔄  Clear Form",     self._clr_stu,  "#888"),
        ])
        tk.Frame(f,bg="#dde3f8",height=2).pack(fill="x",padx=10,pady=2)
        self.st_filters=self._filter_bar(f,[
            ("Course",["All"]+self._course_values,14),
            ("Level",["All"]+LEVELS,12),
            ("Status",["All","Active","Inactive","On Hold"],10),
            ("Fee Status",["All","Paid","Partial","Unpaid"],10),
            ("From",["All"],12),("To",["All"],12),
        ],self._apply_stu_filter,self._clear_stu_filter,
          lambda:self._export_rows("Students",getattr(self,"_filtered_students",[]),STUDENT_HEADERS))
        self.st_tree=self._treef(f,STUDENT_HEADERS)

    def _apply_stu_filter(self):
        rows=self.store.rows("Students")
        f=self.st_filters
        rows=[r for r in rows if (f["Course"].get()=="All" or r.get("Course")==f["Course"].get())]
        rows=[r for r in rows if (f["Level"].get()=="All" or r.get("Level")==f["Level"].get())]
        rows=[r for r in rows if (f["Status"].get()=="All" or r.get("Status")==f["Status"].get())]
        if f["Fee Status"].get()!="All":
            invoices=self.store.rows("Invoices")
            rows=[r for r in rows if any(str(i.get("Student ID"))==str(r.get("Student ID")) and
                  i.get("Payment Status")==f["Fee Status"].get() for i in invoices)]
        rows=[r for r in rows if self._date_ok(r.get("Joining Date",""),f["From"].get(),f["To"].get())]
        self._filtered_students=rows; self._fill(self.st_tree,rows,STUDENT_HEADERS)

    def _clear_stu_filter(self):
        for c in self.st_filters.values(): c.set("All")
        self._apply_stu_filter()

    def _prev_sid(self,_=None):
        c=self.sco.get(); l=self.slv.get()
        if c and l:
            code=COURSE_CODES.get(c,c[:2].upper())
            lvl=l.replace("Level ","L").replace("All Levels","ALL").replace(" ","").upper()
            self._sidlbl.config(text=f"ID preview: RK-{code}-{lvl}-{datetime.now().year}-###")

    def _add_stu(self):
        if not self.sn.get().strip(): return messagebox.showerror("Missing","Student name required")
        if not self.sco.get(): return messagebox.showerror("Missing","Select course")
        if not self.slv.get(): return messagebox.showerror("Missing","Select level")
        if not self.sph.get().strip(): return messagebox.showerror("Missing","Phone required")
        try:
            sid=self.store.add_student({"name":self.sn.get().strip(),"parent":self.sp.get().strip(),
                "phone":self.sph.get().strip(),"email":self.se.get().strip(),
                "grade":self.sg.get().strip(),"course":self.sco.get(),"level":self.slv.get(),
                "batch":self.sba.get(),"joining":self.sj.get(),
                "address":self.sa.get().strip(),"notes":self.sno.get().strip()})
            messagebox.showinfo("Added ✓",f"Student registered!\nID: {sid}")
            self._clr_stu(); self.refresh_all()
        except PermissionError as e: self._pe(str(e))

    def _edit_stu(self):
        sel=self.st_tree.selection()
        if not sel: return messagebox.showwarning("Select","Select a student to edit")
        sid=self.st_tree.item(sel[0],"values")[0]
        data=next((r for r in self.store.rows("Students") if str(r.get("Student ID"))==str(sid)),None)
        if not data: return
        dlg=StudentDlg(self,data)
        if dlg.result:
            try: self.store.edit_student(sid,dlg.result); messagebox.showinfo("Updated ✓","Student updated"); self.refresh_all()
            except PermissionError as e: self._pe(str(e))

    def _del_stu(self):
        sel=self.st_tree.selection()
        if not sel: return messagebox.showwarning("Select","Select a student to delete")
        sid=self.st_tree.item(sel[0],"values")[0]
        name=self.st_tree.item(sel[0],"values")[1]
        if messagebox.askyesno("Confirm Delete",f"Delete student:\n{sid} — {name}\n\nThis cannot be undone!"):
            try: self.store.delete_student(sid); self.refresh_all()
            except PermissionError as e: self._pe(str(e))

    def _clr_stu(self):
        for w in [self.sn,self.sp,self.sph,self.se,self.sg,self.sa,self.sno]: w.delete(0,tk.END)
        self.sco.set(""); self.slv.set(""); self.sba.set("")
        self.sj.delete(0,tk.END); self.sj.insert(0,datetime.now().strftime("%d-%m-%Y"))
        self._sidlbl.config(text="ID preview: select Course & Level")

    # ══════════════════════════════════════════════════════════════════════════
    # INVOICES
    # ══════════════════════════════════════════════════════════════════════════
    def _t_inv(self):
        f=tk.Frame(self.nb,bg=BG_WHITE); self.nb.add(f,text="🧾  Create Invoice")
        slbl(f,"  New Invoice").pack(anchor="w",padx=14,pady=(10,0))
        form=self._ff_frame(f); form.columnconfigure((1,3,5),weight=1)

        tk.Label(form,text="Select Student *",bg=BG_WHITE,fg="#444",
                 font=("Segoe UI",9)).grid(row=0,column=0,sticky="w",padx=(10,4),pady=4)
        self.i_stu=ttk.Combobox(form,width=34,state="readonly")
        self.i_stu.grid(row=0,column=1,columnspan=2,sticky="ew",padx=(0,10),pady=4)
        self.i_stu.bind("<<ComboboxSelected>>",self._fill_stu_inv)
        self.i_idate=self._ff(form,"Invoice Date",0,3,val=datetime.now().strftime("%d-%m-%Y"))
        self.i_due  =self._ff(form,"Due Date",0,5)
        self.i_cr   =self._ff(form,"Course",1,0)
        self.i_lv   =self._ff(form,"Level",1,2)
        self.i_ba   =self._ff(form,"Batch",1,4)
        self.i_fee  =self._ff(form,"Fee (₹) *",2,0)

        # Plan loader
        tk.Label(form,text="Load Plan",bg=BG_WHITE,fg="#444",
                 font=("Segoe UI",9)).grid(row=2,column=2,sticky="w",padx=(10,4),pady=4)
        self.i_plan=ttk.Combobox(form,width=28,state="readonly")
        self.i_plan.grid(row=2,column=3,columnspan=2,sticky="ew",padx=(0,10),pady=4)
        self.i_plan.bind("<<ComboboxSelected>>",self._load_plan)

        # Discount
        tk.Label(form,text="Discount",bg=BG_WHITE,fg="#444",
                 font=("Segoe UI",9)).grid(row=3,column=0,sticky="w",padx=(10,4),pady=4)
        di=tk.Frame(form,bg=BG_WHITE); di.grid(row=3,column=1,sticky="ew",padx=(0,8),pady=4)
        self.i_dv=ttk.Entry(di,width=10); self.i_dv.pack(side="left"); self.i_dv.insert(0,"0")
        self.i_dt=tk.StringVar(value="%")
        tk.Radiobutton(di,text="%",variable=self.i_dt,value="%",bg=BG_WHITE,
                       font=("Segoe UI",9),fg=f"#{PRIMARY}").pack(side="left",padx=2)
        tk.Radiobutton(di,text="₹",variable=self.i_dt,value="₹",bg=BG_WHITE,
                       font=("Segoe UI",9),fg=f"#{PRIMARY}").pack(side="left")

        # Tax
        tk.Label(form,text="Tax (GST)",bg=BG_WHITE,fg="#444",
                 font=("Segoe UI",9)).grid(row=3,column=2,sticky="w",padx=(10,4),pady=4)
        ti=tk.Frame(form,bg=BG_WHITE); ti.grid(row=3,column=3,sticky="ew",padx=(0,8),pady=4)
        self.i_tv=ttk.Entry(ti,width=10); self.i_tv.pack(side="left"); self.i_tv.insert(0,"0")
        self.i_tt=tk.StringVar(value="%")
        tk.Radiobutton(ti,text="%",variable=self.i_tt,value="%",bg=BG_WHITE,
                       font=("Segoe UI",9),fg=f"#{PRIMARY}").pack(side="left",padx=2)
        tk.Radiobutton(ti,text="₹",variable=self.i_tt,value="₹",bg=BG_WHITE,
                       font=("Segoe UI",9),fg=f"#{PRIMARY}").pack(side="left")

        self.i_note=self._ff(form,"Notes",4,0,w=50,cs=5)

        self._tlbl=tk.Label(form,text="",bg=BG_WHITE,font=("Segoe UI",9,"bold"),fg=f"#{ACCENT}")
        self._tlbl.grid(row=5,column=0,columnspan=6,sticky="w",padx=10,pady=2)
        for w in [self.i_fee,self.i_dv,self.i_tv]:
            w.bind("<KeyRelease>",self._prev_total)
        self.i_dt.trace_add("write",lambda *_:self._prev_total())
        self.i_tt.trace_add("write",lambda *_:self._prev_total())

        self._abar(f,[
            ("🧾  Create Invoice",   self._create_inv, f"#{SECONDARY}"),
            ("🖨️  Print Selected",   self._print_inv,  f"#{PRIMARY}"),
            ("🗑️  Delete Selected",  self._del_inv,    f"#{DANGER}"),
            ("🔄  Clear Form",       self._clr_inv,    "#888"),
        ])
        tk.Frame(f,bg="#dde3f8",height=2).pack(fill="x",padx=10,pady=2)
        self.inv_filters=self._filter_bar(f,[
            ("Course",["All"]+self._course_values,14),
            ("Level",["All"]+LEVELS,12),
            ("Status",["All","Paid","Partial","Unpaid"],10),
            ("From",["All"],12),("To",["All"],12),
        ],self._apply_inv_filter,self._clear_inv_filter,
          lambda:self._export_rows("Invoices",getattr(self,"_filtered_invoices",[]),INVOICE_HEADERS))
        self.i_tree=self._treef(f,INVOICE_HEADERS)

    def _apply_inv_filter(self):
        rows=self.store.rows("Invoices")
        f=self.inv_filters
        rows=[r for r in rows if (f["Course"].get()=="All" or r.get("Course")==f["Course"].get())]
        rows=[r for r in rows if (f["Level"].get()=="All" or r.get("Level")==f["Level"].get())]
        rows=[r for r in rows if (f["Status"].get()=="All" or r.get("Payment Status")==f["Status"].get())]
        rows=[r for r in rows if self._date_ok(r.get("Date",""),f["From"].get(),f["To"].get())]
        self._filtered_invoices=rows; self._fill(self.i_tree,rows,INVOICE_HEADERS)

    def _clear_inv_filter(self):
        for c in self.inv_filters.values(): c.set("All")
        self._apply_inv_filter()

    def _fill_stu_inv(self,_=None):
        s=self._slookup.get(self.i_stu.get())
        if not s: return
        for e,k in [(self.i_cr,"Course"),(self.i_lv,"Level"),(self.i_ba,"Batch")]:
            e.config(state="normal"); e.delete(0,tk.END); e.insert(0,s.get(k,""))
        self._reload_plans(s.get("Course",""))

    def _reload_plans(self,course=""):
        plans=self.store.rows("CoursePlans")
        if course: plans=[p for p in plans if p.get("Course")==course or p.get("Course")=="Yearly All-Courses"]
        self._plan_map={
            f"{p.get('Course')} | {p.get('Level')} | {p.get('Duration')} — ₹{money(p.get('Fee',0)):,.0f}":p
            for p in plans if p.get("Active","Yes")=="Yes"}
        self.i_plan["values"]=list(self._plan_map.keys())

    def _load_plan(self,_=None):
        p=self._plan_map.get(self.i_plan.get())
        if not p: return
        self.i_fee.delete(0,tk.END); self.i_fee.insert(0,str(money(p.get("Fee",0))))
        self._prev_total()

    def _prev_total(self,_=None):
        try:
            fee=money(self.i_fee.get()); dv=money(self.i_dv.get()); tv=money(self.i_tv.get())
            da=round(fee*dv/100,2) if self.i_dt.get()=="%" else dv
            sub=fee-da; ta=round(sub*tv/100,2) if self.i_tt.get()=="%" else tv
            tot=sub+ta; bal=tot
            self._tlbl.config(text=
                f"  Fee ₹{fee:,.2f}  –  Disc ₹{da:,.2f}  +  Tax ₹{ta:,.2f}"
                f"  =  Total ₹{tot:,.2f}   |   Balance ₹{bal:,.2f}")
        except: pass

    def _create_inv(self):
        s=self._slookup.get(self.i_stu.get())
        if not s: return messagebox.showerror("Missing","Select a student")
        if not self.i_fee.get().strip(): return messagebox.showerror("Missing","Enter fee")
        try:
            inv_no=self.store.create_invoice({"date":self.i_idate.get(),"sid":s.get("Student ID"),
                "sname":s.get("Student Name"),"course":self.i_cr.get(),"level":self.i_lv.get(),
                "batch":self.i_ba.get(),"fee":self.i_fee.get(),"dt":self.i_dt.get(),
                "dv":self.i_dv.get(),"tt":self.i_tt.get(),"tv":self.i_tv.get(),
                "due":self.i_due.get(),"notes":self.i_note.get()})
            self._open_html(inv_no)
            messagebox.showinfo("Created ✓",f"Invoice: {inv_no}\nOpened for printing.")
            self._clr_inv(); self.refresh_all()
        except PermissionError as e: self._pe(str(e))

    def _print_inv(self):
        sel=self.i_tree.selection()
        inv_no=(self.i_tree.item(sel[0],"values")[0] if sel else
                simpledialog.askstring("Invoice No","Enter Invoice Number:",parent=self))
        if inv_no: self._open_html(str(inv_no))

    def _del_inv(self):
        sel=self.i_tree.selection()
        if not sel: return messagebox.showwarning("Select","Select an invoice to delete")
        inv_no=self.i_tree.item(sel[0],"values")[0]
        if messagebox.askyesno("Confirm",f"Delete invoice {inv_no}?"):
            try: self.store.delete_invoice(inv_no); self.refresh_all()
            except PermissionError as e: self._pe(str(e))

    def _clr_inv(self):
        for w in [self.i_cr,self.i_lv,self.i_ba,self.i_fee,self.i_note,self.i_due]: w.delete(0,tk.END)
        self.i_dv.delete(0,tk.END); self.i_dv.insert(0,"0")
        self.i_tv.delete(0,tk.END); self.i_tv.insert(0,"0")
        self.i_stu.set(""); self.i_plan.set("")
        self.i_dt.set("%"); self.i_tt.set("%"); self._tlbl.config(text="")

    def _open_html(self,inv_no):
        inv=self.store.get_invoice(inv_no)
        if not inv: return messagebox.showerror("Not Found",f"Invoice {inv_no} not found")
        html=make_html(self.store.settings(),inv)
        path=INVOICE_DIR/f"{inv_no}.html"; path.write_text(html,encoding="utf-8")
        webbrowser.open(path.resolve().as_uri())

    # ══════════════════════════════════════════════════════════════════════════
    # PAYMENTS
    # ══════════════════════════════════════════════════════════════════════════
    def _t_pay(self):
        f=tk.Frame(self.nb,bg=BG_WHITE); self.nb.add(f,text="💳  Payments")
        slbl(f,"  Add Payment / Receipt").pack(anchor="w",padx=14,pady=(10,0))
        form=self._ff_frame(f); form.columnconfigure((1,3,5),weight=1)
        self.pi_inv=self._fc(form,"Invoice ID",[],0,0); self.pi_inv.config(state="normal")
        self.pi_sid=self._fc(form,"Student ID",[],0,2); self.pi_sid.config(state="normal")
        self.pam=self._ff(form,"Amount (₹) *",0,4)
        self.pm =self._fc(form,"Payment Mode",PAY_MODES,1,0); self.pm.set("Cash")
        self.prb=self._ff(form,"Received By",1,2,val="Admin")
        self.pnt=self._ff(form,"Notes",1,4)
        detail=tk.Frame(form,bg="#f4f6ff",bd=1,relief="solid")
        detail.grid(row=2,column=0,columnspan=6,sticky="ew",padx=10,pady=6)
        self._pinf=tk.Label(detail,text="Select an invoice ID or student ID to load payment details.",
                            bg="#f4f6ff",font=("Segoe UI",9,"bold"),fg="#444",
                            justify="left",anchor="w")
        self._pinf.pack(fill="x",padx=10,pady=8)
        for w in [self.pi_inv,self.pi_sid]:
            w.bind("<<ComboboxSelected>>",self._lookup_inv)
            w.bind("<FocusOut>",self._lookup_inv)
            w.bind("<Return>",self._lookup_inv)
        self.pi_sid.bind("<KeyRelease>",self._filter_payment_students)
        self._abar(f,[
            ("💳  Add Payment & Print Receipt",self._add_pay,     f"#{SECONDARY}"),
            ("🖨️  Reprint Selected Receipt",   self._reprint,     f"#{PRIMARY}"),
            ("🗑️  Delete Selected Payment",    self._del_pay,     f"#{DANGER}"),
        ])
        tk.Frame(f,bg="#dde3f8",height=2).pack(fill="x",padx=10,pady=2)
        self.pay_filters=self._filter_bar(f,[
            ("Course",["All"]+self._course_values,14),
            ("Level",["All"]+LEVELS,12),
            ("Status",["All","Paid","Partial","Unpaid"],10),
            ("From",["All"],12),("To",["All"],12),
        ],self._apply_pay_filter,self._clear_pay_filter,
          lambda:self._export_rows("Payments",getattr(self,"_filtered_payments",[]),PAYMENT_HEADERS))
        self.p_tree=self._treef(f,PAYMENT_HEADERS)

    def _apply_pay_filter(self):
        rows=self.store.rows("Payments")
        invoices={str(r.get("Invoice No")):r for r in self.store.rows("Invoices")}
        f=self.pay_filters
        def inv_for(r): return invoices.get(str(r.get("Invoice No")),{})
        rows=[r for r in rows if (f["Course"].get()=="All" or inv_for(r).get("Course")==f["Course"].get())]
        rows=[r for r in rows if (f["Level"].get()=="All" or inv_for(r).get("Level")==f["Level"].get())]
        rows=[r for r in rows if (f["Status"].get()=="All" or r.get("Payment Status")==f["Status"].get())]
        rows=[r for r in rows if self._date_ok(r.get("Date",""),f["From"].get(),f["To"].get())]
        self._filtered_payments=rows; self._fill(self.p_tree,rows,PAYMENT_HEADERS)

    def _clear_pay_filter(self):
        for c in self.pay_filters.values(): c.set("All")
        self._apply_pay_filter()

    def _lookup_inv(self,_=None):
        prefer_student=bool(_ and getattr(_,"widget",None) is getattr(self,"pi_sid",None))
        inv_key=self.pi_inv.get().strip()
        sid_key=self._student_id_from_pick(self.pi_sid.get())
        inv=None
        if not prefer_student and inv_key:
            inv=self.store.get_invoice(inv_key)
        if not inv and sid_key:
            inv=self.store.find_invoice_for_payment(sid_key)
        if not inv and inv_key:
            inv=self.store.find_invoice_for_payment(inv_key)
        if inv:
            self.pi_inv.set(str(inv.get("Invoice No","")))
            self.pi_sid.set(f"{inv.get('Student ID','')} - {inv.get('Student Name','')}")
            self._pinf.config(fg=f"#{SECONDARY}",
                text=f"Student Name: {inv.get('Student Name','')}\n"
                     f"Student ID: {inv.get('Student ID','')}    Invoice ID: {inv.get('Invoice No','')}\n"
                     f"Course Joined: {inv.get('Course','')} - {inv.get('Level','')}\n"
                     f"Course Fees: ₹{money(inv.get('Total Amount')):,.2f}    "
                     f"Fees Paid: ₹{money(inv.get('Paid Amount')):,.2f}    "
                     f"Balance: ₹{money(inv.get('Balance')):,.2f}\n"
                     f"Status: {inv.get('Payment Status','')}")
            receipts=[r for r in self.store.rows("Payments")
                      if str(r.get("Invoice No"))==str(inv.get("Invoice No"))]
            self._fill(self.p_tree,receipts,PAYMENT_HEADERS)
        else: self._pinf.config(fg=f"#{DANGER}",text="✗ Invoice / Student ID not found")

    def _add_pay(self):
        inv=self.store.get_invoice(self.pi_inv.get().strip()) or self.store.find_invoice_for_payment(self._student_id_from_pick(self.pi_sid.get()))
        if not inv: return messagebox.showerror("Not Found","Invoice / Student ID not found")
        if not self.pam.get().strip(): return messagebox.showerror("Missing","Enter amount")
        try:
            rno=self.store.add_payment({"inv":inv["Invoice No"],"sid":inv["Student ID"],
                "sname":inv["Student Name"],"mode":self.pm.get(),
                "paid":self.pam.get(),"rcvd":self.prb.get(),"notes":self.pnt.get()})
            upd=self.store.get_invoice(inv["Invoice No"])
            payment=self.store.get_payment(rno)
            html=make_html(self.store.settings(),upd,receipt_no=rno,payment=payment)
            path=INVOICE_DIR/f"{rno}.html"; path.write_text(html,encoding="utf-8")
            webbrowser.open(path.resolve().as_uri())
            messagebox.showinfo("Receipt ✓",f"Receipt: {rno}\nOpened for printing.")
            for w in [self.pi_inv,self.pi_sid,self.pam,self.pnt]: w.delete(0,tk.END)
            self._pinf.config(text=""); self.refresh_all()
        except PermissionError as e: self._pe(str(e))

    def _reprint(self):
        sel=self.p_tree.selection()
        if not sel: return messagebox.showwarning("Select","Select a receipt to reprint")
        vals=self.p_tree.item(sel[0],"values"); rno=vals[0]; inv_no=vals[2]
        inv=self.store.get_invoice(inv_no)
        if not inv: return messagebox.showerror("Not Found","Invoice not found")
        payment=self.store.get_payment(rno)
        html=make_html(self.store.settings(),inv,receipt_no=rno,payment=payment)
        path=INVOICE_DIR/f"{rno}.html"; path.write_text(html,encoding="utf-8")
        webbrowser.open(path.resolve().as_uri())

    def _del_pay(self):
        sel=self.p_tree.selection()
        if not sel: return messagebox.showwarning("Select","Select a payment to delete")
        rno=self.p_tree.item(sel[0],"values")[0]
        if messagebox.askyesno("Confirm",f"Delete receipt {rno}?\nInvoice balance and finance income will auto-adjust."):
            try: self.store.delete_payment(rno); self.refresh_all()
            except PermissionError as e: self._pe(str(e))

    # ══════════════════════════════════════════════════════════════════════════
    # COURSE PLANS
    # ══════════════════════════════════════════════════════════════════════════
    def _t_plans(self):
        f=tk.Frame(self.nb,bg=BG_WHITE); self.nb.add(f,text="📋  Course Plans")
        info=tk.Frame(f,bg="#e8f5e9"); info.pack(fill="x",padx=10,pady=(10,4))
        tk.Label(info,bg="#e8f5e9",fg="#1b5e20",font=("Segoe UI",9),anchor="w",
                 text="  💡 Set standard fees per course & duration. "
                      "Use 'Load Plan' in Create Invoice to auto-fill. "
                      "Yearly All-Courses = student attends ALL courses for the year."
                 ).pack(fill="x",padx=8,pady=6)
        self._abar(f,[
            ("➕  Add Course",      self._add_course, f"#{ACCENT}"),
            ("➕  Add Plan",        self._add_plan,  f"#{SECONDARY}"),
            ("✏️  Edit Selected",   self._edit_plan, f"#{PRIMARY}"),
            ("🗑️  Delete Selected", self._del_plan,  f"#{DANGER}"),
        ])
        self.plan_filters=self._filter_bar(f,[
            ("Course",["All"]+self._course_values,14),
            ("Level",["All"]+LEVELS,12),
            ("Status",["All","Yes","No"],8),
            ("Duration",["All"]+DURATIONS,10),
        ],self._apply_plan_filter,self._clear_plan_filter,
          lambda:self._export_rows("Course_Plans",getattr(self,"_filtered_plans",[]),PLAN_HEADERS))
        self.pl_tree=self._treef(f,PLAN_HEADERS,h=18)

    def _apply_plan_filter(self):
        rows=self.store.rows("CoursePlans")
        f=self.plan_filters
        rows=[r for r in rows if (f["Course"].get()=="All" or r.get("Course")==f["Course"].get())]
        rows=[r for r in rows if (f["Level"].get()=="All" or r.get("Level")==f["Level"].get())]
        rows=[r for r in rows if (f["Status"].get()=="All" or r.get("Active")==f["Status"].get())]
        rows=[r for r in rows if (f["Duration"].get()=="All" or r.get("Duration")==f["Duration"].get())]
        self._filtered_plans=rows; self._fill(self.pl_tree,rows,PLAN_HEADERS)

    def _clear_plan_filter(self):
        for c in self.plan_filters.values(): c.set("All")
        self._apply_plan_filter()

    def _add_course(self):
        name=simpledialog.askstring("Add Course","Enter new course name:",parent=self)
        if not name: return
        try:
            course=self.store.add_course(name)
            messagebox.showinfo("Added ✓",f"Course added: {course}")
            self.refresh_all()
        except ValueError as e:
            messagebox.showerror("Course",str(e))
        except PermissionError as e:
            self._pe(str(e))

    def _add_plan(self):
        dlg=PlanDlg(self)
        if dlg.result:
            try: pid=self.store.add_plan(dlg.result); messagebox.showinfo("Added ✓",f"Plan: {pid}"); self.refresh_all()
            except PermissionError as e: self._pe(str(e))

    def _edit_plan(self):
        sel=self.pl_tree.selection()
        if not sel: return messagebox.showwarning("Select","Select a plan")
        pid=self.pl_tree.item(sel[0],"values")[0]
        data=next((r for r in self.store.rows("CoursePlans") if str(r.get("Plan ID"))==str(pid)),None)
        if not data: return
        dlg=PlanDlg(self,data)
        if dlg.result:
            try: self.store.edit_plan(pid,dlg.result); messagebox.showinfo("Updated ✓","Plan updated"); self.refresh_all()
            except PermissionError as e: self._pe(str(e))

    def _del_plan(self):
        sel=self.pl_tree.selection()
        if not sel: return messagebox.showwarning("Select","Select a plan")
        pid=self.pl_tree.item(sel[0],"values")[0]
        if messagebox.askyesno("Confirm",f"Delete plan {pid}?"):
            try: self.store.delete_plan(pid); self.refresh_all()
            except PermissionError as e: self._pe(str(e))

    # ══════════════════════════════════════════════════════════════════════════
    # FINANCE TRACKER
    # ══════════════════════════════════════════════════════════════════════════
    def _t_finance(self):
        f=tk.Frame(self.nb,bg=BG_WHITE); self.nb.add(f,text="💰  Finance")

        # Summary cards
        top=tk.Frame(f,bg=BG_WHITE); top.pack(fill="x",padx=14,pady=(12,4))
        self._fc_cards={}
        for key,col,bg in [
            ("Total Income",  f"#{SUCCESS}", "#f0fdf4"),
            ("Total Expenses",f"#{DANGER}",  "#fff7f7"),
            ("Net Profit",    "#7c3aed",     "#faf5ff"),
            ("This Month Income", f"#{SUCCESS}","#f0fdf4"),
            ("This Month Expenses",f"#{DANGER}", "#fff7f7"),
        ]:
            c=tk.Frame(top,bg=bg,width=148,height=78,relief="flat")
            c.pack(side="left",padx=6,pady=4,expand=True,fill="both"); c.pack_propagate(False)
            tk.Label(c,text=key,bg=bg,font=("Segoe UI",7),fg="#666",wraplength=120).pack(pady=(10,1))
            lbl=tk.Label(c,text="—",bg=bg,font=("Segoe UI",13,"bold"),fg=col)
            lbl.pack(); self._fc_cards[key]=lbl

        # Filter bar
        fbar=tk.Frame(f,bg=BG_WHITE); fbar.pack(fill="x",padx=12,pady=(4,2))
        tk.Label(fbar,text="Filter:",bg=BG_WHITE,font=("Segoe UI",9)).pack(side="left")
        self._fin_type=ttk.Combobox(fbar,values=["All","Income","Expense"],
                                    state="readonly",width=10); self._fin_type.set("All")
        self._fin_type.pack(side="left",padx=6)
        self._fin_month=ttk.Combobox(fbar,values=["All"]+MONTHS,state="readonly",width=8)
        self._fin_month.set("All"); self._fin_month.pack(side="left",padx=4)
        yr=datetime.now().year
        self._fin_year=ttk.Combobox(fbar,values=["All"]+[str(y) for y in range(yr-2,yr+3)],
                                    state="readonly",width=8)
        self._fin_year.set(str(yr)); self._fin_year.pack(side="left",padx=4)
        tk.Label(fbar,text="From:",bg=BG_WHITE,font=("Segoe UI",9)).pack(side="left",padx=(6,2))
        self._fin_from=ttk.Combobox(fbar,values=["All"],state="normal",width=12)
        self._fin_from.set("All"); self._fin_from.pack(side="left",padx=4)
        tk.Label(fbar,text="To:",bg=BG_WHITE,font=("Segoe UI",9)).pack(side="left",padx=(6,2))
        self._fin_to=ttk.Combobox(fbar,values=["All"],state="normal",width=12)
        self._fin_to.set("All"); self._fin_to.pack(side="left",padx=4)
        sbtn(fbar,"🔍 Apply Filter",self._apply_fin_filter,bg=f"#{PRIMARY}",px=10,py=5).pack(side="left",padx=6)
        sbtn(fbar,"🔄 Clear",self._clear_fin_filter,bg="#888",px=10,py=5).pack(side="left")
        sbtn(fbar,"⬇ Export Excel",lambda:self._export_rows("Finance",getattr(self,"_filtered_finance",[]),FINANCE_HEADERS),
             bg=f"#{SECONDARY}",px=10,py=5).pack(side="left",padx=4)

        self._abar(f,[
            ("➕  Add Income",     lambda:self._add_fin("Income"),  f"#{SUCCESS}"),
            ("➕  Add Expense",    lambda:self._add_fin("Expense"), f"#{DANGER}"),
            ("✏️  Edit Selected",  self._edit_fin,                  f"#{PRIMARY}"),
            ("🗑️  Delete Selected",self._del_fin,                   f"#{DANGER}"),
        ])
        tk.Frame(f,bg="#dde3f8",height=2).pack(fill="x",padx=10,pady=2)

        # Colour legend
        leg=tk.Frame(f,bg=BG_WHITE); leg.pack(fill="x",padx=12,pady=(2,4))
        tk.Label(leg,text="■ Income",bg=BG_WHITE,fg=f"#{SUCCESS}",
                 font=("Segoe UI",8)).pack(side="left",padx=8)
        tk.Label(leg,text="■ Expense",bg=BG_WHITE,fg=f"#{DANGER}",
                 font=("Segoe UI",8)).pack(side="left",padx=4)

        self.fin_tree=self._treef(f,FINANCE_HEADERS,h=11)

    def _add_fin(self,dtype="Income"):
        dlg=FinanceDlg(self,default_type=dtype)
        if dlg.result:
            try: eid=self.store.add_finance(dlg.result); messagebox.showinfo("Added ✓",f"Entry: {eid}"); self.refresh_all()
            except PermissionError as e: self._pe(str(e))

    def _edit_fin(self):
        sel=self.fin_tree.selection()
        if not sel: return messagebox.showwarning("Select","Select an entry to edit")
        eid=self.fin_tree.item(sel[0],"values")[0]
        data=next((r for r in self.store.rows("Finance") if str(r.get("Entry ID"))==str(eid)),None)
        if not data: return
        dlg=FinanceDlg(self,data)
        if dlg.result:
            try: self.store.edit_finance(eid,dlg.result); messagebox.showinfo("Updated ✓","Entry updated"); self.refresh_all()
            except PermissionError as e: self._pe(str(e))

    def _del_fin(self):
        sel=self.fin_tree.selection()
        if not sel: return messagebox.showwarning("Select","Select an entry to delete")
        eid=self.fin_tree.item(sel[0],"values")[0]
        if messagebox.askyesno("Confirm",f"Delete entry {eid}?"):
            try: self.store.delete_finance(eid); self.refresh_all()
            except PermissionError as e: self._pe(str(e))

    def _apply_fin_filter(self):
        rows=self.store.rows("Finance")
        ft=self._fin_type.get(); fm=self._fin_month.get(); fy=self._fin_year.get()
        if ft!="All": rows=[r for r in rows if r.get("Type")==ft]
        if fm!="All":
            mi=MONTHS.index(fm)+1
            rows=[r for r in rows if self._date_month(r.get("Date",""))==mi]
        if fy!="All":
            rows=[r for r in rows if self._date_year(r.get("Date",""))==int(fy)]
        rows=[r for r in rows if self._date_ok(r.get("Date",""),self._fin_from.get(),self._fin_to.get())]
        self._filtered_finance=rows
        self._fill_fin(self.fin_tree,rows,FINANCE_HEADERS)

    def _clear_fin_filter(self):
        self._fin_type.set("All"); self._fin_month.set("All")
        self._fin_year.set(str(datetime.now().year))
        self._fin_from.set("All"); self._fin_to.set("All")
        self._apply_fin_filter()

    @staticmethod
    def _date_month(ds):
        try:
            for fmt in ["%d-%m-%Y","%Y-%m-%d","%d/%m/%Y"]:
                try: return datetime.strptime(str(ds),fmt).month
                except: pass
        except: pass
        return 0

    @staticmethod
    def _date_year(ds):
        try:
            for fmt in ["%d-%m-%Y","%Y-%m-%d","%d/%m/%Y"]:
                try: return datetime.strptime(str(ds),fmt).year
                except: pass
        except: pass
        return 0

    # ══════════════════════════════════════════════════════════════════════════
    # REPORTS
    # ══════════════════════════════════════════════════════════════════════════
    def _t_reports(self):
        f=tk.Frame(self.nb,bg=BG_WHITE); self.nb.add(f,text="📁  Reports")
        slbl(f,"  Quick Actions").pack(anchor="w",padx=14,pady=(14,6))
        bar=tk.Frame(f,bg=BG_WHITE); bar.pack(fill="x",padx=14,pady=4)
        for txt,cmd,bg in [
            ("🔄 Refresh All",   self.refresh_all,  f"#{SECONDARY}"),
            ("📂 Open Excel",    lambda:webbrowser.open(DATA_FILE.resolve().as_uri()),f"#{PRIMARY}"),
            ("🗂️ Invoices Folder",lambda:webbrowser.open(INVOICE_DIR.resolve().as_uri()),f"#{PRIMARY}"),
            ("📁 Switch File",   self._switch_file, "#555"),
        ]: sbtn(bar,txt,cmd,bg=bg).pack(side="left",padx=5,pady=4)
        tk.Frame(f,bg="#dde3f8",height=2).pack(fill="x",padx=14,pady=10)
        slbl(f,"  Invoice Reports").pack(anchor="w",padx=14,pady=(0,4))
        pc=["Invoice No","Date","Student Name","Total Amount","Paid Amount","Balance","Due Date","Payment Status"]
        self.rpt_cols=pc
        self.rpt_filters=self._filter_bar(f,[
            ("Course",["All"]+self._course_values,14),
            ("Level",["All"]+LEVELS,12),
            ("Status",["All","Paid","Partial","Unpaid"],10),
            ("From",["All"],12),("To",["All"],12),
        ],self._apply_report_filter,self._clear_report_filter,
          lambda:self._export_rows("Reports",getattr(self,"_filtered_reports",[]),self.rpt_cols))
        self.rpt_tree=self._treef(f,pc,h=13)
        tk.Label(f,text="  💡 Backup robokidy_billing_data.xlsx regularly.",
                 bg=BG_WHITE,fg="#888",font=("Segoe UI",9,"italic")).pack(pady=6)

    def _switch_file(self):
        sel=filedialog.asksaveasfilename(defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],title="Choose / Create Data File",
            initialdir=str(BASE_DIR))
        if sel:
            global DATA_FILE
            DATA_FILE=Path(sel); self.store=ExcelStore(DATA_FILE); self.refresh_all()

    def _apply_report_filter(self):
        rows=self.store.rows("Invoices")
        f=self.rpt_filters
        rows=[r for r in rows if (f["Course"].get()=="All" or r.get("Course")==f["Course"].get())]
        rows=[r for r in rows if (f["Level"].get()=="All" or r.get("Level")==f["Level"].get())]
        rows=[r for r in rows if (f["Status"].get()=="All" or r.get("Payment Status")==f["Status"].get())]
        rows=[r for r in rows if self._date_ok(r.get("Date",""),f["From"].get(),f["To"].get())]
        self._filtered_reports=rows
        self._fill(self.rpt_tree,rows,self.rpt_cols)

    def _clear_report_filter(self):
        for c in self.rpt_filters.values(): c.set("All")
        self._apply_report_filter()

    # ══════════════════════════════════════════════════════════════════════════
    # REFRESH ALL
    # ══════════════════════════════════════════════════════════════════════════
    def refresh_all(self):
        try: self.store.update_dashboard()
        except PermissionError: pass

        students=self.store.rows("Students")
        invoices=self.store.rows("Invoices")
        payments=self.store.rows("Payments")
        plans   =self.store.rows("CoursePlans")
        finance =self.store.rows("Finance")
        self._course_values=self.store.course_values()

        tb=sum(money(r.get("Total Amount")) for r in invoices)
        tr=sum(money(r.get("Paid Amount"))  for r in invoices)
        tbl=sum(money(r.get("Balance"))     for r in invoices)
        inc=sum(money(r.get("Amount")) for r in finance if r.get("Type")=="Income")
        exp=sum(money(r.get("Amount")) for r in finance if r.get("Type")=="Expense")

        now=datetime.now()
        m_inc=sum(money(r.get("Amount")) for r in finance
                  if r.get("Type")=="Income" and self._date_month(r.get("Date",""))==now.month
                  and self._date_year(r.get("Date",""))==now.year)
        m_exp=sum(money(r.get("Amount")) for r in finance
                  if r.get("Type")=="Expense" and self._date_month(r.get("Date",""))==now.month
                  and self._date_year(r.get("Date",""))==now.year)

        for k,v in [("Students",str(len(students))),("Invoices",str(len(invoices))),
                    ("Receipts",str(len(payments))),("Billing",f"₹{tb:,.0f}"),
                    ("Received",f"₹{tr:,.0f}"),("Balance",f"₹{tbl:,.0f}"),
                    ("Income",f"₹{inc:,.0f}"),("Expenses",f"₹{exp:,.0f}"),
                    ("Net Profit",f"₹{inc-exp:,.0f}")]:
            if k in self._dc: self._dc[k].config(text=v)

        for k,v in [("Total Income",f"₹{inc:,.2f}"),("Total Expenses",f"₹{exp:,.2f}"),
                    ("Net Profit",f"₹{inc-exp:,.2f}"),
                    ("This Month Income",f"₹{m_inc:,.2f}"),
                    ("This Month Expenses",f"₹{m_exp:,.2f}")]:
            if k in self._fc_cards: self._fc_cards[k].config(text=v)

        self._fill(self.st_tree, students, STUDENT_HEADERS)
        self._fill(self.i_tree,  invoices, INVOICE_HEADERS)
        self._fill(self.p_tree,  payments, PAYMENT_HEADERS)
        self._fill(self.pl_tree, plans,    PLAN_HEADERS)
        self._fill_fin(self.fin_tree, finance, FINANCE_HEADERS)

        sh=["Invoice No","Date","Student Name","Total Amount","Paid Amount","Balance","Payment Status"]
        self._fill(self._dt, invoices[-10:][::-1], sh)

        pending=[r for r in invoices if r.get("Payment Status")!="Paid"]
        pc=["Invoice No","Date","Student Name","Total Amount","Paid Amount","Balance","Due Date","Payment Status"]
        self._fill(self.rpt_tree, pending, pc)

        self._slookup={
            f"{s.get('Student ID')} — {s.get('Student Name')} ({s.get('Course','')})":s
            for s in students}
        if hasattr(self,"sco"): self.sco["values"]=self._course_values
        for attr in ["st_filters","inv_filters","pay_filters","plan_filters","rpt_filters"]:
            flt=getattr(self,attr,None)
            if flt and "Course" in flt:
                cur=flt["Course"].get()
                flt["Course"]["values"]=["All"]+self._course_values
                flt["Course"].set(cur if cur in flt["Course"]["values"] else "All")
        if hasattr(self,"st_filters"):
            dates=self._dates_for(students,"Joining Date")
            self.st_filters["From"]["values"]=dates; self.st_filters["To"]["values"]=dates
            self._apply_stu_filter()
        if hasattr(self,"inv_filters"):
            dates=self._dates_for(invoices,"Date")
            self.inv_filters["From"]["values"]=dates; self.inv_filters["To"]["values"]=dates
            self._apply_inv_filter()
        if hasattr(self,"pay_filters"):
            dates=self._dates_for(payments,"Date")
            self.pay_filters["From"]["values"]=dates; self.pay_filters["To"]["values"]=dates
            self._apply_pay_filter()
        if hasattr(self,"plan_filters"):
            self._apply_plan_filter()
        if hasattr(self,"_fin_from"):
            dates=self._dates_for(finance,"Date")
            self._fin_from["values"]=dates; self._fin_to["values"]=dates
            self._apply_fin_filter()
        if hasattr(self,"rpt_filters"):
            dates=self._dates_for(invoices,"Date")
            self.rpt_filters["From"]["values"]=dates; self.rpt_filters["To"]["values"]=dates
            self._apply_report_filter()
        if hasattr(self,"i_stu"): self.i_stu["values"]=list(self._slookup.keys())
        if hasattr(self,"pi_inv"):
            self.pi_inv["values"]=[str(r.get("Invoice No","")) for r in invoices if r.get("Invoice No")]
        if hasattr(self,"pi_sid"):
            seen=[]
            student_names={str(s.get("Student ID","")):str(s.get("Student Name","")) for s in students}
            for r in invoices:
                sid=str(r.get("Student ID",""))
                name=str(r.get("Student Name") or student_names.get(sid,""))
                label=f"{sid} - {name}" if name else sid
                if sid and label not in seen:
                    seen.append(label)
            self._payment_student_values=seen
            self.pi_sid["values"]=seen
        self._reload_plans()

if __name__=="__main__":
    app=BillingApp(); app.mainloop()
